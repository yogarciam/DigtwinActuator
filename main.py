import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# Configurar Times New Roman en matplotlib
plt.rcParams["font.family"] = "Times New Roman"

# ---------------------------------------------------------------------------
# Función auxiliar para anexar unidades a las etiquetas de los ejes
# ---------------------------------------------------------------------------
def with_units(label):
    lower = label.lower()
    if lower == "time":
        return f"{label} (s)"
    elif lower in ["laser experimental", "unity theoretical"]:
        return f"{label} (m)"
    elif lower == "force":
        return f"{label} (N)"
    else:
        return label

# ---------------------------------------------------------------------------
# 1. Configuración de rutas (raw string para Windows)
# ---------------------------------------------------------------------------
ROOT_DIR = r"C:\Users\YOVANI\Downloads\Frailejon\2025_1\Polytech\Final_comparition\DatosConsolidados"  # Ajusta la ruta principal
OUTPUT_DIR = os.path.join(ROOT_DIR, "graficos_resultados")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# 2. Mapeo de frecuencia basado en el nombre del archivo
# ---------------------------------------------------------------------------
frecuencia_map = {
    "frequency_1": "0.25Hz",
    "frequency_2": "0.125Hz",
    "frequency_3": "0.05Hz"
}

# ---------------------------------------------------------------------------
# 3. Función para leer y limpiar datos de la hoja "Promedios"
# ---------------------------------------------------------------------------
def leer_promedios(path_excel, frecuencia):
    """
    Lee la hoja 'Promedios' de un archivo Excel (.xlsm) utilizando la primera
    fila como cabecera, renombra columnas duplicadas y realiza las siguientes
    conversiones:
      - Renombra "slide" a "Unity theoretical".
      - Renombra "Experimental" a "Laser experimental".
      - Crea la columna "Force" a partir de "Pressure" usando:
            Force = 0.00079173 * 6894.76 * Pressure
    Además, agrega la columna 'frecuencia'.
    """
    wb = load_workbook(path_excel, data_only=True)
    if "Promedios" not in wb.sheetnames:
        return None

    data = list(wb["Promedios"].values)
    # Filtrar filas completamente vacías
    data = [row for row in data if any(cell is not None for cell in row)]
    if len(data) < 2:
        return None

    columns = list(data[0])
    data_rows = data[1:]
    col_count = {}
    final_cols = []
    for col in columns:
        if col is None:
            col = "Unnamed"
        if col not in col_count:
            col_count[col] = 0
            final_cols.append(col)
        else:
            col_count[col] += 1
            final_cols.append(f"{col}_{col_count[col]}")

    df = pd.DataFrame(data_rows, columns=final_cols)
    df["frecuencia"] = frecuencia

    # Renombrar columnas: primera letra mayúscula, resto en minúsculas
    if "slide" in df.columns:
        df.rename(columns={"slide": "Unity theoretical"}, inplace=True)
    if "Experimental" in df.columns:
        df.rename(columns={"Experimental": "Laser experimental"}, inplace=True)

    # Calcular la columna Force a partir de Pressure
    if "Pressure" in df.columns:
        df["Force"] = 0.00079173 * 6894.76 * df["Pressure"]

    return df

# ---------------------------------------------------------------------------
# 4. Función genérica de graficación en escala de grises
# ---------------------------------------------------------------------------
def graficar(df_all, x, y, nombre, titulo, carpeta_salida, markersize=2, linewidth=0.8):
    plt.style.use('grayscale')
    plt.figure(figsize=(10, 5))
    markers = ['o', 's', '^', 'd', 'v', '*', 'p']

    for i, (freq, df_freq) in enumerate(df_all.groupby("frecuencia")):
        marker = markers[i % len(markers)]
        # Se usa la etiqueta original (sin unidad) para la leyenda.
        label = f"{y[0].upper() + y[1:].lower()} ({freq})"
        plt.plot(df_freq[x], df_freq[y],
                 label=label,
                 linestyle='-',
                 marker=marker,
                 markersize=markersize,
                 linewidth=linewidth)

    plt.xlabel(with_units(x), fontsize=12)
    plt.ylabel(with_units(y), fontsize=12)
    plt.title(titulo, fontsize=14)
    plt.legend(fontsize=10)
    # Forzar que los ejes comiencen en 0
    plt.xlim(left=0)
    plt.ylim(bottom=0)
    plt.tight_layout()
    plt.savefig(os.path.join(carpeta_salida, f"{nombre}.png"), dpi=300)
    plt.savefig(os.path.join(carpeta_salida, f"{nombre}.pdf"), dpi=300)
    plt.close()

# ---------------------------------------------------------------------------
# 5. Función para graficar Laser experimental vs Force (con ejes invertidos)
# ---------------------------------------------------------------------------
def graficar_laser_vs_force(df_all, carpeta_salida):
    """
    Grafica, con ejes invertidos:
      - Eje x: Laser experimental (m)
      - Eje y: Force (N)
    Se utiliza un markersize reducido para hacer la gráfica más legible.
    """
    plt.style.use('grayscale')
    plt.figure(figsize=(10, 5))
    markers = ['o', 's', '^', 'd', 'v', '*', 'p']

    for i, (freq, df_freq) in enumerate(df_all.groupby("frecuencia")):
        marker = markers[i % len(markers)]
        plt.plot(df_freq["Laser experimental"], df_freq["Force"],
                 label=f"Laser experimental ({freq})",
                 linestyle='-',
                 marker=marker,
                 markersize=3,  # ligeramente mayor para visibilidad pero reducido
                 linewidth=0.8)

    plt.xlabel(with_units("Laser experimental"), fontsize=12)
    plt.ylabel(with_units("Force"), fontsize=12)
    plt.title("Laser experimental vs force", fontsize=14)
    plt.legend(fontsize=10)
    plt.xlim(left=0)
    plt.ylim(bottom=0)
    plt.tight_layout()
    plt.savefig(os.path.join(carpeta_salida, "laser_vs_force.png"), dpi=300)
    plt.savefig(os.path.join(carpeta_salida, "laser_vs_force.pdf"), dpi=300)
    plt.close()

# ---------------------------------------------------------------------------
# 6. Función para graficar comparativa: time vs laser experimental y unity theoretical
# ---------------------------------------------------------------------------
def graficar_experimental_vs_theoretical(df_all, carpeta_salida):
    plt.style.use('grayscale')
    plt.figure(figsize=(10, 5))
    markers = ['o', 's', '^', 'd', 'v', '*', 'p']

    for i, (freq, df_freq) in enumerate(df_all.groupby("frecuencia")):
        marker = markers[i % len(markers)]
        if {"time", "Laser experimental", "Unity theoretical"}.issubset(df_freq.columns):
            plt.plot(df_freq["time"], df_freq["Laser experimental"],
                     label=f"Laser experimental ({freq})",
                     linestyle="--",
                     marker=marker,
                     markersize=2,
                     linewidth=0.8)
            plt.plot(df_freq["time"], df_freq["Unity theoretical"],
                     label=f"Unity theoretical ({freq})",
                     linestyle=":",
                     marker=marker,
                     markersize=2,
                     linewidth=0.8)

    plt.xlabel(with_units("time"), fontsize=12)
    plt.ylabel("Measurements", fontsize=12)
    plt.title("Time vs laser experimental & unity theoretical", fontsize=14)
    plt.legend(fontsize=10)
    plt.xlim(left=0)
    plt.ylim(bottom=0)
    plt.tight_layout()
    plt.savefig(os.path.join(carpeta_salida, "time_vs_experimental_vs_theoretical.png"), dpi=300)
    plt.savefig(os.path.join(carpeta_salida, "time_vs_experimental_vs_theoretical.pdf"), dpi=300)
    plt.close()

# ---------------------------------------------------------------------------
# 7. Procesamiento por carpeta (cada carpeta se procesa de forma independiente)
# ---------------------------------------------------------------------------
for folder in os.listdir(ROOT_DIR):
    if not folder.startswith("Consolidate_final_results_"):
        continue

    folder_path = os.path.join(ROOT_DIR, folder)
    proc_path = os.path.join(folder_path, "Processing")
    if not os.path.exists(proc_path):
        continue

    # Carpeta de salida para la carpeta actual
    carpeta_salida = os.path.join(OUTPUT_DIR, folder)
    os.makedirs(carpeta_salida, exist_ok=True)

    consolidado_total = []
    for archivo in os.listdir(proc_path):
        if archivo.startswith("frequency_") and archivo.endswith(".xlsm"):
            clave = archivo.replace(".xlsm", "")
            frecuencia = frecuencia_map.get(clave)
            path_excel = os.path.join(proc_path, archivo)
            df = leer_promedios(path_excel, frecuencia)
            if df is not None and not df.empty:
                df.to_csv(os.path.join(carpeta_salida, f"{frecuencia}_promedios.csv"), index=False)
                consolidado_total.append(df)

    if consolidado_total:
        df_todo = pd.concat(consolidado_total, ignore_index=True)
        df_todo.to_csv(os.path.join(carpeta_salida, "consolidado_total.csv"), index=False)

        # Gráficas con etiquetas y ejes con unidades, forzando a iniciar en 0
        if {"Unity theoretical", "Laser experimental"}.issubset(df_todo.columns):
            graficar(df_todo, "Unity theoretical", "Laser experimental",
                     "unity_vs_laser",
                     "Unity theoretical vs laser experimental",
                     carpeta_salida)
        if {"time", "Unity theoretical"}.issubset(df_todo.columns):
            graficar(df_todo, "time", "Unity theoretical",
                     "time_vs_unity",
                     "Time vs unity theoretical",
                     carpeta_salida)
        if {"time", "Laser experimental"}.issubset(df_todo.columns):
            graficar(df_todo, "time", "Laser experimental",
                     "time_vs_laser",
                     "Time vs laser experimental",
                     carpeta_salida)
        if {"time", "dac_bits"}.issubset(df_todo.columns):
            graficar(df_todo, "time", "dac_bits",
                     "time_vs_dac_bits",
                     "Time vs dac bits",
                     carpeta_salida)
        if {"Laser experimental", "Force"}.issubset(df_todo.columns):
            graficar_laser_vs_force(df_todo, carpeta_salida)
        graficar_experimental_vs_theoretical(df_todo, carpeta_salida)

        print(f"✅ Análisis completado para la carpeta: {folder}")
    else:
        print(f"❌ No se encontraron datos válidos para la carpeta: {folder}")
